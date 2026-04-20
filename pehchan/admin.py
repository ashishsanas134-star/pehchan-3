from django.contrib import admin
from django.urls import path, reverse
from django.utils.html import format_html
from django.utils import timezone
from django import forms
from django.db import models
import csv
from django.http import HttpResponse
from .models import (
    Event, EventVolunteer, LifetimeVolunteer, 
    Certificate, MaterialDonation, MoneyDonation, DonorCertificate, ContactMessage, AnonymousDonation,
    PehchanWallet, WalletTransaction, Expense
)


class ExportDataMixin:
    def export_as_csv(self, request, queryset):
        meta = self.model._meta
        field_names = [field.name for field in meta.fields]

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename={meta}.csv'
        writer = csv.writer(response)

        writer.writerow(field_names)
        for obj in queryset:
            row = []
            for field in field_names:
                try:
                    value = getattr(obj, field)
                    if value and hasattr(value, 'url'):
                        value = value.url
                    elif hasattr(value, 'pk'):
                        value = str(value)
                    else:
                        value = str(value) if value is not None else ""
                except Exception:
                    value = "ERROR"
                row.append(value)
            writer.writerow(row)

        return response

    def export_as_excel(self, request, queryset):
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        meta = self.model._meta
        field_names = [field.name for field in meta.fields]
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={meta}.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(meta).split('.')[-1].capitalize()
        
        # Write headers with style
        header_font = Font(bold=True)
        for col_num, column_title in enumerate(field_names, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title.replace('_', ' ').capitalize()
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field in enumerate(field_names, 1):
                cell = ws.cell(row=row_num, column=col_num)
                try:
                    value = getattr(obj, field)
                    
                    if value and hasattr(value, 'url'):
                        # Handle image/file fields with clickable links
                        cell.value = "View File"
                        cell.hyperlink = value.url
                        cell.font = Font(color="0000FF", underline="single")
                    elif hasattr(value, 'pk'):
                        cell.value = str(value)
                    else:
                        cell.value = str(value) if value is not None else ""
                except Exception:
                    cell.value = "ERROR"
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    val_str = str(cell.value)
                    if len(val_str) > max_length:
                        max_length = len(val_str)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 50)
            
        wb.save(response)
        return response

    export_as_csv.short_description = "Download selected as CSV"
    export_as_excel.short_description = "Download selected as Excel (XLSX)"


class NativeDateTimeWidget(forms.SplitDateTimeWidget):
    def __init__(self, attrs=None, date_format=None, time_format=None):
        widgets = (
            forms.DateInput(attrs={'type': 'date', 'class': 'vDateField'}, format=date_format),
            forms.TimeInput(attrs={'type': 'time', 'class': 'vTimeField'}, format=time_format),
        )
        # Note: we don't call super().__init__ because we want to define our own widgets
        forms.MultiWidget.__init__(self, widgets, attrs)



@admin.register(Event)
class EventAdmin(admin.ModelAdmin, ExportDataMixin):
    list_display = ('name', 'event_date', 'get_status_display', 'status', 'created_by')
    readonly_fields = ('created_by',)
    actions = ['export_as_csv', 'export_as_excel']
    
    def save_model(self, request, obj, form, change):
        if not obj.pk:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)

    class Media:
        css = {
            'all': ('CSS/admin_patch.css',)
        }

    def formfield_for_dbfield(self, db_field, request, **kwargs):
        if isinstance(db_field, models.DateTimeField):
            kwargs['widget'] = NativeDateTimeWidget
        return super().formfield_for_dbfield(db_field, request, **kwargs)


@admin.register(EventVolunteer)
class EventVolunteerAdmin(admin.ModelAdmin, ExportDataMixin):
    list_display = ['user', 'event', 'contact_number', 'joined_at', 'status', 'issue_certificate_button']
    list_filter = ['status', 'joined_at', 'event']
    search_fields = ['user__username', 'user__email', 'event__name', 'contact_number']
    date_hierarchy = 'joined_at'
    ordering = ['-joined_at']
    actions = ['issue_certificate', 'export_as_csv', 'export_as_excel']
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('issue-certificate/<int:volunteer_id>/', self.admin_site.admin_view(self.issue_certificate_view), name='pehchan_eventvolunteer_issue_certificate'),
        ]
        return custom_urls + urls
    
    def issue_certificate_view(self, request, volunteer_id):
        """Directly render the certificate issuance form within the admin"""
        from django.shortcuts import render, get_object_or_404
        from django.contrib import messages
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        from .models import EventVolunteer, Certificate
        
        # Check permissions
        if not request.user.is_staff:
            messages.error(request, 'You do not have permission to access this page.')
            return HttpResponseRedirect(reverse('admin:pehchan_eventvolunteer_changelist'))
        
        volunteer = get_object_or_404(EventVolunteer, id=volunteer_id)
        
        # Check if certificate already exists
        existing_certificate = Certificate.objects.filter(event=volunteer.event, volunteer=volunteer).first()
        if existing_certificate:
            messages.error(request, 'A certificate already exists for this volunteer.')
            return HttpResponseRedirect(reverse('admin:pehchan_eventvolunteer_changelist'))
        
        if request.method == 'POST':
            certificate_file = request.FILES.get('certificate_file')
            try:
                # Create a certificate - number will be auto-generated by model's save()
                from django.utils import timezone
                certificate = Certificate.objects.create(
                    event=volunteer.event,
                    volunteer=volunteer,
                    issue_date=timezone.now().date()
                )
                # If a file was uploaded, save it to the certificate
                if certificate_file:
                    certificate.file = certificate_file
                    certificate.save()
                messages.success(request, f'Certificate issued successfully for {volunteer.user.username}.')
                return HttpResponseRedirect(reverse('admin:pehchan_eventvolunteer_changelist'))
            except Exception as e:
                messages.error(request, f'Error issuing certificate: {str(e)}')
        
        context = {
            'volunteer': volunteer,
            'certificate_type': 'Volunteer',
            'title': 'Issue Volunteer Certificate',
        }
        
        return render(request, 'admin/issue_certificate.html', context)
    
    def issue_certificate_button(self, obj):
        """Display an 'Issue Certificate' button for each volunteer"""
        from django.urls import reverse
        # Only show button for approved volunteers
        if obj.status == 'approved':
            url = reverse('admin:pehchan_eventvolunteer_issue_certificate', args=[obj.pk])
            return format_html(
                '<a class="button" href="{}" target="_blank">Issue Certificate</a>',
                url
            )
        return "Not Approved"
    issue_certificate_button.short_description = 'Certificate Actions'
    
    def issue_certificate(self, request, queryset):
        """Issue certificates for selected event volunteers"""
        from .models import Certificate
        from django.utils import timezone
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        
        # If only one volunteer is selected, redirect to custom form for certificate number input
        if len(queryset) == 1:
            volunteer = queryset.first()
            if volunteer.status == 'approved':
                # Redirect to a custom view for entering certificate number
                url = reverse('admin:pehchan_eventvolunteer_issue_certificate', args=[volunteer.pk])
                return HttpResponseRedirect(url)
            else:
                self.message_user(request, 'Only approved volunteers can be issued certificates.', level='error')
                return
        
        # For multiple selection, issue certificates with auto-generated numbers
        created_count = 0
        skipped_count = 0
        not_approved_count = 0
        
        for volunteer in queryset:
            if volunteer.status != 'approved':
                not_approved_count += 1
                continue
            
            # Check if certificate already exists
            if not Certificate.objects.filter(event=volunteer.event, volunteer=volunteer).exists():
                Certificate.objects.create(
                    event=volunteer.event,
                    volunteer=volunteer,
                    issue_date=timezone.now().date()
                )
                created_count += 1
            else:
                skipped_count += 1
        
        if created_count > 0:
            self.message_user(request, f'Successfully issued {created_count} certificate(s).')
        if skipped_count > 0:
            self.message_user(request, f'Skipped {skipped_count} volunteer(s) who already have certificates.', level='warning')
        if not_approved_count > 0:
            self.message_user(request, f'Skipped {not_approved_count} volunteer(s) who are not approved.', level='warning')
    
    issue_certificate.short_description = 'Issue Certificate'


@admin.register(LifetimeVolunteer)
class LifetimeVolunteerAdmin(admin.ModelAdmin, ExportDataMixin):
    list_display = ['user', 'contact_number', 'joined_at', 'status', 'approved_at', 'rejected_at']
    list_filter = ['status', 'joined_at']
    search_fields = ['user__username', 'user__email', 'contact_number', 'motivation']
    date_hierarchy = 'joined_at'
    ordering = ['-joined_at']
    readonly_fields = ['joined_at', 'approved_at', 'rejected_at', 'reviewed_by']
    actions = ['approve_volunteers', 'reject_volunteers', 'export_as_csv', 'export_as_excel']
    
    def approve_volunteers(self, request, queryset):
        """Approve selected lifetime volunteers"""
        updated_count = 0
        for volunteer in queryset:
            if volunteer.status != 'approved':
                volunteer.status = 'approved'
                volunteer.approved_at = timezone.now()
                volunteer.rejected_at = None
                volunteer.reviewed_by = request.user
                volunteer.save()
                updated_count += 1
        
        if updated_count > 0:
            self.message_user(request, f'Successfully approved {updated_count} lifetime volunteer(s).')
        else:
            self.message_user(request, 'No volunteers were updated.', level='warning')
    
    approve_volunteers.short_description = 'Approve selected lifetime volunteers'
    
    def reject_volunteers(self, request, queryset):
        """Reject selected lifetime volunteers"""
        updated_count = 0
        for volunteer in queryset:
            if volunteer.status != 'rejected':
                volunteer.status = 'rejected'
                volunteer.rejected_at = timezone.now()
                volunteer.approved_at = None
                volunteer.reviewed_by = request.user
                volunteer.save()
                updated_count += 1
        
        if updated_count > 0:
            self.message_user(request, f'Successfully rejected {updated_count} lifetime volunteer(s).')
        else:
            self.message_user(request, 'No volunteers were updated.', level='warning')
    
    reject_volunteers.short_description = 'Reject selected lifetime volunteers'


@admin.register(Certificate)
class CertificateAdmin(admin.ModelAdmin, ExportDataMixin):
    list_display = ['certificate_number', 'volunteer', 'event', 'issue_date']
    list_filter = ['issue_date', 'event']
    search_fields = ['certificate_number', 'volunteer__user__username', 'event__name']
    date_hierarchy = 'issue_date'
    ordering = ['-issue_date']
    actions = ['export_as_csv', 'export_as_excel']


@admin.register(MaterialDonation)
class MaterialDonationAdmin(admin.ModelAdmin, ExportDataMixin):
    list_display = ['user', 'item_name', 'quantity', 'location', 'contact_number', 'status', 'created_at', 'issue_certificate_button']
    list_filter = ['status', 'location', 'created_at']
    search_fields = ['user__username', 'item_name', 'location', 'contact_number']
    date_hierarchy = 'created_at'
    ordering = ['-created_at']
    readonly_fields = ['created_at']
    actions = ['issue_donor_certificate', 'export_as_csv', 'export_as_excel']
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('issue-certificate/<int:donation_id>/', self.admin_site.admin_view(self.issue_certificate_view), name='pehchan_materialdonation_issue_certificate'),
        ]
        return custom_urls + urls
    
    def issue_certificate_view(self, request, donation_id):
        """Directly render the certificate issuance form within the admin"""
        from django.shortcuts import render, get_object_or_404
        from django.contrib import messages
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        from .models import MaterialDonation, DonorCertificate
        from django.db import IntegrityError
        
        # Check permissions
        if not request.user.is_staff:
            messages.error(request, 'You do not have permission to access this page.')
            return HttpResponseRedirect(reverse('admin:pehchan_materialdonation_changelist'))
        
        donation = get_object_or_404(MaterialDonation, id=donation_id)
        
        # Check if certificate already exists
        existing_certificate = DonorCertificate.objects.filter(material_donation=donation).first()
        if existing_certificate:
            messages.error(request, 'A certificate already exists for this donation.')
            return HttpResponseRedirect(reverse('admin:pehchan_materialdonation_changelist'))
        
        if request.method == 'POST':
            certificate_file = request.FILES.get('certificate_file')
            try:
                donor_certificate = DonorCertificate.objects.create(
                    user=donation.user,
                    donation_type='material',
                    material_donation=donation,
                    issue_date=timezone.now().date(),
                    remarks=f'Thank you for your generous donation of {donation.quantity} {donation.item_name}(s).'
                )
                # If a file was uploaded, save it to the certificate
                if certificate_file:
                    donor_certificate.file = certificate_file
                    donor_certificate.save()
                messages.success(request, f'Donor certificate issued successfully for {donation.user.username}.')
                return HttpResponseRedirect(reverse('admin:pehchan_materialdonation_changelist'))
            except Exception as e:
                messages.error(request, f'Error issuing certificate: {str(e)}')
        
        context = {
            'donation': donation,
            'certificate_type': 'Donor (Material)',
            'title': 'Issue Material Donation Certificate',
        }
        
        return render(request, 'admin/issue_certificate.html', context)
    
    def issue_certificate_button(self, obj):
        """Display an 'Issue Certificate' button for each donation"""
        from django.urls import reverse
        # Only show button for received donations
        if obj.status == 'received':
            url = reverse('admin:pehchan_materialdonation_issue_certificate', args=[obj.pk])
            return format_html(
                '<a class="button" href="{}" target="_blank">Issue Certificate</a>',
                url
            )
        return "Not Received"
    issue_certificate_button.short_description = 'Certificate Actions'
    
    def issue_donor_certificate(self, request, queryset):
        """Issue donor certificates for selected material donations"""
        from .models import DonorCertificate
        from django.utils import timezone
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        
        # If only one donation is selected, redirect to custom form for certificate number input
        if len(queryset) == 1:
            donation = queryset.first()
            if donation.status == 'received':
                # Redirect to a custom view for entering certificate number
                url = reverse('admin:pehchan_materialdonation_issue_certificate', args=[donation.pk])
                return HttpResponseRedirect(url)
            else:
                self.message_user(request, 'Only received donations can be issued certificates.', level='error')
                return
        
        # For multiple selection, issue certificates with auto-generated numbers
        created_count = 0
        skipped_count = 0
        not_received_count = 0
        
        for donation in queryset:
            if donation.status != 'received':
                not_received_count += 1
                continue
            
            # Check if donor certificate already exists
            if not DonorCertificate.objects.filter(material_donation=donation).exists():
                try:
                    DonorCertificate.objects.create(
                        user=donation.user,
                        donation_type='material',
                        material_donation=donation,
                        issue_date=timezone.now().date(),
                        remarks=f'Thank you for your generous donation of {donation.quantity} {donation.item_name}(s).'
                    )
                    created_count += 1
                except IntegrityError:
                    skipped_count += 1
            else:
                skipped_count += 1
        
        if created_count > 0:
            self.message_user(request, f'Successfully issued {created_count} donor certificate(s).')
        if skipped_count > 0:
            self.message_user(request, f'Skipped {skipped_count} donation(s) that already have certificates.', level='warning')
        if not_received_count > 0:
            self.message_user(request, f'Skipped {not_received_count} donation(s) that are not received.', level='warning')
    
    issue_donor_certificate.short_description = 'Issue Donor Certificate'


@admin.register(MoneyDonation)
class MoneyDonationAdmin(admin.ModelAdmin, ExportDataMixin):
    list_display = ['user', 'amount', 'contact_number', 'status', 'payment_method', 'created_at', 'issue_certificate_button']
    list_filter = ['status', 'payment_method', 'created_at']
    search_fields = ['user__username', 'user__email', 'upi_id', 'contact_number']
    date_hierarchy = 'created_at'
    ordering = ['-created_at']
    readonly_fields = ['created_at']
    actions = ['issue_donor_certificate', 'verify_donations', 'decline_donations', 'export_as_csv', 'export_as_excel']
    
    def verify_donations(self, request, queryset):
        queryset.update(status='verified')
    verify_donations.short_description = "Mark selected donations as Verified"

    def decline_donations(self, request, queryset):
        queryset.update(status='declined')
    decline_donations.short_description = "Mark selected donations as Declined"
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('issue-certificate/<int:donation_id>/', self.admin_site.admin_view(self.issue_certificate_view), name='pehchan_moneydonation_issue_certificate'),
        ]
        return custom_urls + urls
    
    def issue_certificate_view(self, request, donation_id):
        """Directly render the certificate issuance form within the admin"""
        from django.shortcuts import render, get_object_or_404
        from django.contrib import messages
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        from .models import MoneyDonation, DonorCertificate
        from django.db import IntegrityError
        
        # Check permissions
        if not request.user.is_staff:
            messages.error(request, 'You do not have permission to access this page.')
            return HttpResponseRedirect(reverse('admin:pehchan_moneydonation_changelist'))
        
        donation = get_object_or_404(MoneyDonation, id=donation_id)
        
        # Check if certificate already exists
        existing_certificate = DonorCertificate.objects.filter(money_donation=donation).first()
        if existing_certificate:
            messages.error(request, 'A certificate already exists for this donation.')
            return HttpResponseRedirect(reverse('admin:pehchan_moneydonation_changelist'))
        
        if request.method == 'POST':
            certificate_file = request.FILES.get('certificate_file')
            try:
                donor_certificate = DonorCertificate.objects.create(
                    user=donation.user,
                    donation_type='money',
                    money_donation=donation,
                    issue_date=timezone.now().date(),
                    remarks=f'Thank you for your generous donation of Rs. {donation.amount}.'
                )
                # If a file was uploaded, save it to the certificate
                if certificate_file:
                    donor_certificate.file = certificate_file
                    donor_certificate.save()
                messages.success(request, f'Donor certificate issued successfully for {donation.user.username}.')
                return HttpResponseRedirect(reverse('admin:pehchan_moneydonation_changelist'))
            except Exception as e:
                messages.error(request, f'Error issuing certificate: {str(e)}')
        
        context = {
            'donation': donation,
            'certificate_type': 'Donor (Money)',
            'title': 'Issue Money Donation Certificate',
        }
        
        return render(request, 'admin/issue_certificate.html', context)
    
    def issue_certificate_button(self, obj):
        """Display an 'Issue Certificate' button for each donation"""
        from django.urls import reverse
        # Only show button for verified donations
        if obj.status == 'verified':
            url = reverse('admin:pehchan_moneydonation_issue_certificate', args=[obj.pk])
            return format_html(
                '<a class="button" href="{}" target="_blank">Issue Certificate</a>',
                url
            )
        return format_html('<span style="color: #999;">Requires Verification</span>')
    issue_certificate_button.short_description = 'Certificate Actions'
    
    def issue_donor_certificate(self, request, queryset):
        """Issue donor certificates for selected money donations"""
        from .models import DonorCertificate
        from django.utils import timezone
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        
        # Filter for verified donations only
        verified_queryset = queryset.filter(status='verified')
        unverified_count = queryset.count() - verified_queryset.count()
        
        if unverified_count > 0:
            self.message_user(request, f'Skipped {unverified_count} donation(s) that are not verified.', level='warning')

        if not verified_queryset.exists():
            return
            
        # If only one verified donation is selected, redirect to custom form
        if verified_queryset.count() == 1:
            donation = verified_queryset.first()
            url = reverse('admin:pehchan_moneydonation_issue_certificate', args=[donation.pk])
            return HttpResponseRedirect(url)
        
        # For multiple selection
        created_count = 0
        skipped_count = 0
        
        for donation in verified_queryset:
            if not DonorCertificate.objects.filter(money_donation=donation).exists():
                try:
                    DonorCertificate.objects.create(
                        user=donation.user,
                        donation_type='money',
                        money_donation=donation,
                        issue_date=timezone.now().date(),
                        remarks=f'Thank you for your generous donation of Rs. {donation.amount}.'
                    )
                    created_count += 1
                except IntegrityError:
                    skipped_count += 1
            else:
                skipped_count += 1
        
        if created_count > 0:
            self.message_user(request, f'Successfully issued {created_count} donor certificate(s).')
        if skipped_count > 0:
            self.message_user(request, f'Skipped {skipped_count} donation(s) that already have certificates.', level='warning')
    
    issue_donor_certificate.short_description = 'Issue Donor Certificate'


@admin.register(DonorCertificate)
class DonorCertificateAdmin(admin.ModelAdmin, ExportDataMixin):
    list_display = ['certificate_number', 'user', 'donation_type', 'get_donation_info', 'issue_date', 'issued_by', 'file_link', 'created_at']
    list_filter = ['donation_type', 'issue_date', 'created_at']
    search_fields = ['certificate_number', 'user__username', 'user__email', 'user__first_name', 'user__last_name']
    date_hierarchy = 'issue_date'
    ordering = ['-issue_date', '-created_at']
    readonly_fields = ['certificate_number', 'created_at', 'updated_at', 'get_donation_preview', 'file_link']
    actions = ['export_as_csv', 'export_as_excel']
    
    fieldsets = (
        ('Donor Information', {
            'fields': ('user', 'donation_type')
        }),
        ('Donation Details', {
            'fields': ('material_donation', 'money_donation', 'get_donation_preview')
        }),
        ('Certificate Information', {
            'fields': ('certificate_number', 'issue_date', 'issued_by', 'remarks', 'file')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('user', 'material_donation', 'money_donation')

    def get_donation_info(self, obj):
        """Display donation details in list view"""
        return obj.get_donation_details()
    get_donation_info.short_description = 'Donation Details'
    
    def get_donation_preview(self, obj):
        """Show donation preview in admin form"""
        try:
            if obj.donation_type == 'material' and obj.material_donation:
                donation = obj.material_donation
                return f"Item: {donation.item_name} | Quantity: {donation.quantity} | Location: {donation.location} | Date: {donation.created_at.strftime('%Y-%m-%d')}"
            elif obj.donation_type == 'money' and obj.money_donation:
                donation = obj.money_donation
                return f"Amount: Rs. {donation.amount} | Payment Method: {donation.get_payment_method_display()} | Date: {donation.created_at.strftime('%Y-%m-%d')}"
        except Exception:
            return "Error retrieving donation preview"
        return "No donation selected"
    get_donation_preview.short_description = 'Donation Preview'
    
    def file_link(self, obj):
        if obj.file:
            return format_html('<a href="{}" target="_blank">View Certificate</a>', obj.file.url)
        return "No File"
    file_link.short_description = 'Certificate File'

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """Filter donation options based on donation_type"""
        if db_field.name == "material_donation":
            kwargs["queryset"] = MaterialDonation.objects.select_related('user').all()
        elif db_field.name == "money_donation":
            kwargs["queryset"] = MoneyDonation.objects.select_related('user').all()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


@admin.register(ContactMessage)
class ContactMessageAdmin(admin.ModelAdmin, ExportDataMixin):
    list_display = [
        'colored_status', 
        'name', 
        'email', 
        'short_message_display', 
        'priority_badge',
        'created_at',
        'responded_by',
        'action_buttons'
    ]
    
    list_filter = [
        'status',
        'priority', 
        'created_at',
        'responded_at',
        'responded_by',
        ('created_at', admin.DateFieldListFilter),
    ]
    
    search_fields = [
        'name', 
        'email', 
        'message', 
        'admin_notes',
        'response'
    ]
    
    date_hierarchy = 'created_at'
    
    ordering = ['-created_at']
    
    readonly_fields = [
        'created_at', 
        'updated_at', 
        'ip_address', 
        'user_agent',
        'responded_at'
    ]
    
    fieldsets = (
        ('Contact Information', {
            'fields': ('name', 'email')
        }),
        ('Message Details', {
            'fields': ('message', 'priority', 'status')
        }),
        ('Response', {
            'fields': ('response', 'responded_by', 'responded_at'),
            'classes': ('collapse',)
        }),
        ('Admin Notes', {
            'fields': ('admin_notes',),
            'classes': ('collapse',),
            'description': 'Internal notes - not visible to the sender'
        }),
        ('Metadata', {
            'fields': ('ip_address', 'user_agent', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    actions = [
        'mark_as_read',
        'mark_as_replied', 
        'mark_as_archived',
        'set_priority_high',
        'set_priority_medium',
        'set_priority_low',
        'send_reply',
        'export_as_csv', 'export_as_excel'
    ]
    
    # Specify the custom template
    change_form_template = 'admin/pehchan/contactmessage/change_form.html'
    
    # Custom display methods
    def colored_status(self, obj):
        """Display status with color coding"""
        colors = {
            'new': '#ff5722',
            'read': '#2196f3',
            'replied': '#4caf50',
            'archived': '#9e9e9e',
        }
        bg_colors = {
            'new': '#ffebee',
            'read': '#e3f2fd',
            'replied': '#e8f5e9',
            'archived': '#fafafa',
        }
        color = colors.get(obj.status, '#000')
        bg_color = bg_colors.get(obj.status, '#f5f5f5')
        return format_html(
            '<span style="color: {}; background-color: {}; padding: 3px 8px; border-radius: 3px; font-weight: bold; border: 1px solid {};">{}</span>',
            color,
            bg_color,
            color,
            obj.get_status_display()
        )
    colored_status.short_description = 'Status'
    colored_status.admin_order_field = 'status'
    
    def priority_badge(self, obj):
        """Display priority with badge style"""
        colors = {
            'low': '#4caf50',
            'medium': '#ff9800',
            'high': '#ff5722',
            'urgent': '#e91e63',
        }
        bg_colors = {
            'low': '#e8f5e9',
            'medium': '#fff3e0',
            'high': '#ffebee',
            'urgent': '#fce4ec',
        }
        color = colors.get(obj.priority, '#000')
        bg_color = bg_colors.get(obj.priority, '#f5f5f5')
        return format_html(
            '<span style="background-color: {}; color: {}; padding: 3px 10px; border-radius: 12px; font-size: 11px; font-weight: bold; border: 1px solid {};">{}</span>',
            bg_color,
            color,
            color,
            obj.get_priority_display().upper()
        )
    priority_badge.short_description = 'Priority'
    priority_badge.admin_order_field = 'priority'
    
    def short_message_display(self, obj):
        """Display shortened message"""
        return obj.get_short_message(60)
    short_message_display.short_description = 'Message Preview'
    
    def action_buttons(self, obj):
        """Display quick action buttons"""
        if obj.status == 'new':
            return format_html(
                '<span style="color: #ff5722; font-weight: bold; background: #ffebee; padding: 3px 8px; border-radius: 3px;">⚠ NEW</span>'
            )
        elif obj.status == 'replied':
            return format_html(
                '<span style="color: #4caf50; font-weight: bold; background: #e8f5e9; padding: 3px 8px; border-radius: 3px;">✓ Replied</span>'
            )
        return format_html('<span style="color: #9e9e9e; background: #fafafa; padding: 3px 8px; border-radius: 3px;">-</span>')
    action_buttons.short_description = 'Quick Status'
    
    # Custom actions
    def mark_as_read(self, request, queryset):
        """Mark selected messages as read"""
        updated = queryset.filter(status='new').update(status='read')
        self.message_user(request, f'{updated} message(s) marked as read.')
    mark_as_read.short_description = '📖 Mark as Read'
    
    def mark_as_replied(self, request, queryset):
        """Mark selected messages as replied"""
        updated = queryset.exclude(status='replied').update(
            status='replied',
            responded_at=timezone.now()
        )
        self.message_user(request, f'{updated} message(s) marked as replied.')
    mark_as_replied.short_description = '✉️ Mark as Replied'
    
    def mark_as_archived(self, request, queryset):
        """Archive selected messages"""
        updated = queryset.update(status='archived')
        self.message_user(request, f'{updated} message(s) archived.')
    mark_as_archived.short_description = '📦 Archive Messages'
    
    def set_priority_high(self, request, queryset):
        """Set priority to high"""
        updated = queryset.update(priority='high')
        self.message_user(request, f'{updated} message(s) marked as HIGH priority.')
    set_priority_high.short_description = '🔴 Set Priority: HIGH'
    
    def set_priority_medium(self, request, queryset):
        """Set priority to medium"""
        updated = queryset.update(priority='medium')
        self.message_user(request, f'{updated} message(s) marked as MEDIUM priority.')
    set_priority_medium.short_description = '🟠 Set Priority: MEDIUM'
    
    def set_priority_low(self, request, queryset):
        """Set priority to low"""
        updated = queryset.update(priority='low')
        self.message_user(request, f'{updated} message(s) marked as LOW priority.')
    set_priority_low.short_description = '🟢 Set Priority: LOW'
    
    def send_reply(self, request, queryset):
        """Send reply to selected messages"""
        from django.core.mail import send_mail
        from django.conf import settings
        
        replied_count = 0
        error_count = 0
        
        for message in queryset:
            if not message.response:  # Skip if no response is set
                error_count += 1
                continue
                
            try:
                # Send email to the sender
                send_mail(
                    subject=f'Re: {message.name} - Contact Form Message',
                    message=message.response,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[message.email],
                    fail_silently=False,
                )
                
                # Update message status
                message.status = 'replied'
                message.responded_by = request.user
                message.responded_at = timezone.now()
                message.save()
                
                replied_count += 1
            except Exception as e:
                error_count += 1
                self.message_user(request, f'Failed to send reply to {message.email}: {str(e)}', level='error')
        
        if replied_count > 0:
            self.message_user(request, f'Successfully sent {replied_count} reply(ies).')
        if error_count > 0:
            self.message_user(request, f'Failed to send {error_count} reply(ies).', level='error')
    send_reply.short_description = '📤 Send Reply via Email'
    
    # Override save to auto-set responded_by and send email if reply_message is provided
    def response_add(self, request, obj, post_url_continue=None):
        """Override response_add to handle reply functionality"""
        from django.core.mail import send_mail
        from django.conf import settings
        
        # Check if reply_message was provided
        reply_message = request.POST.get('reply_message')
        if reply_message:
            # Set the response field to the reply message
            obj.response = reply_message
            # Send email to the sender
            try:
                send_mail(
                    subject=f'Re: {obj.name} - Contact Form Message',
                    message=reply_message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[obj.email],
                    fail_silently=False,
                )
                self.message_user(request, f'Reply sent successfully to {obj.email}.')
            except Exception as e:
                self.message_user(request, f'Failed to send reply to {obj.email}: {str(e)}', level='error')
        
        # Original save logic
        if obj.response:
            if not obj.responded_by:
                obj.responded_by = request.user
            if not obj.responded_at:
                obj.responded_at = timezone.now()
            if obj.status == 'new' or obj.status == 'read':
                obj.status = 'replied'
        return super().response_add(request, obj, post_url_continue)
    
    def response_change(self, request, obj):
        """Override response_change to handle reply functionality"""
        from django.core.mail import send_mail
        from django.conf import settings
        
        # Check if reply_message was provided
        reply_message = request.POST.get('reply_message')
        if reply_message:
            # Set the response field to the reply message
            obj.response = reply_message
            # Send email to the sender
            try:
                send_mail(
                    subject=f'Re: {obj.name} - Contact Form Message',
                    message=reply_message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[obj.email],
                    fail_silently=False,
                )
                self.message_user(request, f'Reply sent successfully to {obj.email}.')
            except Exception as e:
                self.message_user(request, f'Failed to send reply to {obj.email}: {str(e)}', level='error')
        
        # Original save logic
        if obj.response:
            if not obj.responded_by:
                obj.responded_by = request.user
            if not obj.responded_at:
                obj.responded_at = timezone.now()
            if obj.status == 'new' or obj.status == 'read':
                obj.status = 'replied'
        return super().response_change(request, obj)


@admin.register(PehchanWallet)
class PehchanWalletAdmin(admin.ModelAdmin, ExportDataMixin):
    list_display = ['balance', 'created_at', 'updated_at']
    readonly_fields = ['created_at', 'updated_at']
    actions = ['export_as_csv', 'export_as_excel']
    
    class Media:
        css = {
            'all': ('CSS/admin_patch.css',)
        }
    
    def has_add_permission(self, request):
        # Only allow one wallet instance
        return not PehchanWallet.objects.exists()
    
    def has_delete_permission(self, request, obj=None):
        # Prevent deletion of the wallet
        return False


@admin.register(WalletTransaction)
class WalletTransactionAdmin(admin.ModelAdmin, ExportDataMixin):
    list_display = ['transaction_type', 'amount', 'description', 'balance_after_transaction', 'created_at']
    list_filter = ['transaction_type', 'created_at']
    search_fields = ['description']
    readonly_fields = ['wallet', 'amount', 'transaction_type', 'description', 'related_donation', 'related_expense', 'balance_after_transaction', 'created_at']
    date_hierarchy = 'created_at'
    actions = ['export_as_csv', 'export_as_excel']
    
    def has_add_permission(self, request):
        # Transactions are created automatically, not manually
        return False
    
    def has_change_permission(self, request, obj=None):
        # Transactions should not be modified
        return False


@admin.register(Expense)
class ExpenseAdmin(admin.ModelAdmin, ExportDataMixin):
    list_display = ['title', 'category', 'amount', 'date', 'approved_by', 'created_at']
    list_filter = ['category', 'date', 'approved_by']
    search_fields = ['title', 'description']
    readonly_fields = ['created_at', 'updated_at']
    actions = ['export_as_csv', 'export_as_excel']
    
    class Media:
        css = {
            'all': ('CSS/admin_patch.css',)
        }
    
    def formfield_for_dbfield(self, db_field, request, **kwargs):
        if isinstance(db_field, (models.DateTimeField, models.TimeField)):
            kwargs['widget'] = NativeDateTimeWidget
        return super().formfield_for_dbfield(db_field, request, **kwargs)
    
    fieldsets = (
        ('Expense Details', {
            'fields': ('title', 'description', 'category', 'amount', 'date')
        }),
        ('Documentation', {
            'fields': ('receipt',)
        }),
        ('Approval', {
            'fields': ('approved_by', 'approved_at')
        }),
        ('Metadata', {
            'fields': ('created_by', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )


@admin.register(AnonymousDonation)
class AnonymousDonationAdmin(admin.ModelAdmin, ExportDataMixin):
    """Admin interface for anonymous donations"""
    list_display = [
        'get_donor_name',
        'amount',
        'transaction_id',
        'has_receipt',
        'wants_receipt',
        'google_form_submitted',
        'created_at',
    ]
    
    list_filter = [
        'wants_receipt',
        'google_form_submitted',
        'created_at',
        ('created_at', admin.DateFieldListFilter),
    ]
    
    search_fields = [
        'donor_name',
        'donor_email',
        'donor_phone',
        'transaction_id',
        'message',
    ]
    
    readonly_fields = ['created_at', 'ip_address']
    actions = ['export_as_csv', 'export_as_excel']
    
    class Media:
        css = {
            'all': ('CSS/admin_patch.css',)
        }
    
    fieldsets = (
        ('Donor Information (Optional)', {
            'fields': ('donor_name', 'donor_email', 'donor_phone')
        }),
        ('Donation Details', {
            'fields': ('amount', 'transaction_id', 'message')
        }),
        ('Receipt & Form', {
            'fields': ('receipt', 'wants_receipt', 'google_form_submitted')
        }),
        ('Metadata', {
            'fields': ('ip_address', 'created_at'),
            'classes': ('collapse',)
        }),
    )
    
    date_hierarchy = 'created_at'
    ordering = ['-created_at']
    
    def get_donor_name(self, obj):
        """Display donor name or Anonymous"""
        return obj.donor_name if obj.donor_name else 'Anonymous'
    get_donor_name.short_description = 'Donor Name'
    get_donor_name.admin_order_field = 'donor_name'
    
    def has_receipt(self, obj):
        """Display receipt status"""
        return "Yes" if obj.receipt else "No"
    has_receipt.short_description = 'Receipt Uploaded'
    has_receipt.admin_order_field = 'receipt'
